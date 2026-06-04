"""
IFC Export Views

Export views for Raumbuch, WoFlV, GAEB, X83 formats.
"""

from django.http import HttpResponse
from django.shortcuts import get_object_or_404
from django.views.generic import View

from .models import IFCModel, Room
from .services.ifc_query_service import extract_ifc_data


class ExportRaumbuchView(View):
    """Raumbuch als Excel exportieren"""

    def get(self, request, model_id):
        ifc_model = get_object_or_404(IFCModel, pk=model_id)
        export_type = request.GET.get("type", "raumbuch")

        from apps.export.services.export_service import RaumbuchExportService

        service = RaumbuchExportService()

        if export_type == "din277":
            output = service.export_din277_summary(ifc_model)
            filename = f"DIN277_{ifc_model.project.name}_v{ifc_model.version}.xlsx"
        else:
            output = service.export_to_excel(ifc_model)
            filename = f"Raumbuch_{ifc_model.project.name}_v{ifc_model.version}.xlsx"

        response = HttpResponse(
            output.read(),
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
        response["Content-Disposition"] = f'attachment; filename="{filename}"'

        return response


class ExportWoFlVView(View):
    """WoFlV Wohnflächenberechnung als Excel exportieren"""

    def get(self, request, model_id):
        from io import BytesIO

        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill

        ifc_model = get_object_or_404(IFCModel, pk=model_id)

        # WoFlV berechnen
        from nl2cad.areas.woflv import WoFlVCalculator

        ifc_data = extract_ifc_data(ifc_model)
        rooms = ifc_data["rooms"]

        calculator = WoFlVCalculator()
        result = calculator.calculate_from_rooms(rooms)

        # Excel erstellen
        wb = Workbook()
        ws = wb.active
        ws.title = "WoFlV Berechnung"

        # Header
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="2E7D32", end_color="2E7D32", fill_type="solid")

        ws["A1"] = "WoFlV Wohnflächenberechnung"
        ws["A1"].font = Font(bold=True, size=14)
        ws["A2"] = ifc_model.project.name

        # Zusammenfassung
        ws["A4"] = "Zusammenfassung"
        ws["A4"].font = Font(bold=True, size=12)

        anrechnungsquote = (
            result.total_woflv_m2 / result.total_raw_m2 * 100 if result.total_raw_m2 > 0 else 0.0
        )
        summary_data = [
            ("Grundfläche gesamt:", result.total_raw_m2, "m²"),
            ("", "", ""),
            ("WOHNFLÄCHE GESAMT:", result.total_woflv_m2, "m²"),
            ("Anrechnungsquote:", anrechnungsquote, "%"),
        ]

        for idx, (label, value, unit) in enumerate(summary_data, 5):
            ws.cell(row=idx, column=1, value=label)
            if value:
                ws.cell(row=idx, column=2, value=value).number_format = "#,##0.00"
            ws.cell(row=idx, column=3, value=unit)

        # Raumdetails
        ws["A12"] = "Raumdetails"
        ws["A12"].font = Font(bold=True, size=12)

        headers = ["Raumname", "Grundfläche", "Höhe", "Faktor", "Wohnfläche"]
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=13, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill

        for row_idx, room in enumerate(result.rooms, 14):
            ws.cell(row=row_idx, column=1, value=room.name)
            ws.cell(
                row=row_idx, column=2, value=round(room.raw_area_m2, 2)
            ).number_format = "#,##0.00"
            ws.cell(row=row_idx, column=3, value=round(room.height_m, 2)).number_format = "0.00"
            ws.cell(row=row_idx, column=4, value=room.factor).number_format = "0%"
            ws.cell(
                row=row_idx, column=5, value=round(room.woflv_area_m2, 2)
            ).number_format = "#,##0.00"

        # Spaltenbreiten
        ws.column_dimensions["A"].width = 20
        ws.column_dimensions["B"].width = 25

        output = BytesIO()
        wb.save(output)
        output.seek(0)

        filename = f"WoFlV_{ifc_model.project.name}_v{ifc_model.version}.xlsx"
        response = HttpResponse(
            output.read(),
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
        response["Content-Disposition"] = f'attachment; filename="{filename}"'
        return response


class ExportGAEBView(View):
    """GAEB Leistungsverzeichnis exportieren"""

    def get(self, request, model_id):
        from decimal import Decimal

        from nl2cad.gaeb.generator import GAEBGenerator

        ifc_model = get_object_or_404(IFCModel, pk=model_id)
        from nl2cad.gaeb.models import Leistungsverzeichnis, LosGruppe, Position

        format_type = request.GET.get("format", "excel")  # excel oder xml

        # Räume laden via Service
        ifc_data = extract_ifc_data(ifc_model)
        rooms_qs = ifc_data["rooms"]

        # LV erstellen
        lv = Leistungsverzeichnis(
            projekt_name=ifc_model.project.name,
            projekt_nummer=str(ifc_model.project.pk)[:8],
        )

        # Los 1: Bodenbeläge
        boden_positionen = [
            Position(
                oz=f"01.{i + 1:03d}",
                kurztext=f"Bodenbelag {r['name']}",
                menge=Decimal(str(round(r["area"] or 0, 2))),
                einheit="m²",
            )
            for i, r in enumerate(rooms_qs)
        ]
        lv.lose.append(LosGruppe(oz="01", bezeichnung="Bodenbeläge", positionen=boden_positionen))

        # Los 2: Sockelleisten
        sockel_positionen = [
            Position(
                oz=f"02.{i + 1:03d}",
                kurztext=f"Sockelleiste {r['name']}",
                menge=Decimal(str(round(r["perimeter"] or 0, 2))),
                einheit="m",
            )
            for i, r in enumerate(rooms_qs)
            if r.get("perimeter")
        ]
        if sockel_positionen:
            lv.lose.append(
                LosGruppe(oz="02", bezeichnung="Sockelleisten", positionen=sockel_positionen)
            )

        # Export
        generator = GAEBGenerator()

        if format_type == "xml":
            output = generator.generate_xml(lv)
            content_type = "application/xml"
            filename = f"LV_{ifc_model.project.name}_v{ifc_model.version}.x84"
        else:
            output = generator.generate_excel(lv)
            content_type = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            filename = f"LV_{ifc_model.project.name}_v{ifc_model.version}.xlsx"

        response = HttpResponse(output.read(), content_type=content_type)
        response["Content-Disposition"] = f'attachment; filename="{filename}"'
        return response


class ExportX83View(View):
    """
    IFC → GAEB X83 Export (Angebot mit Mengen und Preisen)

    Extrahiert alle Mengen aus dem IFC-Modell und erstellt
    ein vollständiges Leistungsverzeichnis nach GAEB X83.

    Query-Parameter:
        format: xml (default) oder excel
        gewerke: kommaseparierte Liste (z.B. bodenbelag,tueren,fenster)
        prices: 1/0 - Einheitspreise inkludieren
    """

    def get(self, request, model_id):
        ifc_model = get_object_or_404(IFCModel, pk=model_id)

        format_type = request.GET.get("format", "xml")
        include_prices = request.GET.get("prices", "1") == "1"

        from nl2cad.gaeb.converter import IFCX83Converter

        # IFC-Daten aus Datenbank laden
        ifc_data = extract_ifc_data(ifc_model)

        # Konvertieren
        converter = IFCX83Converter()

        if format_type == "excel":
            output = converter.convert_to_excel(
                ifc_data=ifc_data,
                projekt_name=ifc_model.project.name,
                projekt_nummer=str(ifc_model.project.pk)[:8],
                include_prices=include_prices,
            )
            content_type = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            filename = f"LV_X83_{ifc_model.project.name}_v{ifc_model.version}.xlsx"
        else:
            output = converter.convert_to_x83(
                ifc_data=ifc_data,
                projekt_name=ifc_model.project.name,
                projekt_nummer=str(ifc_model.project.pk)[:8],
                include_prices=include_prices,
            )
            content_type = "application/xml"
            filename = f"LV_{ifc_model.project.name}_v{ifc_model.version}.x83"

        response = HttpResponse(output.read(), content_type=content_type)
        response["Content-Disposition"] = f'attachment; filename="{filename}"'
        return response
